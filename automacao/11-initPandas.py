import pandas as pd
from openpyxl.workbook import Workbook

lista = [
    ['coluna', 'nome', 'telefone', 'email'],
    ['1', 'Wilson', '(11)00000-0000', 'wilson@teste100.com'],
    ['2', 'Jose da Silva', '(11) 11111-1111', 'jose@teste100.com.br'],
    ['3', 'Maria luiza', '(11) 22222-2222', 'maria@teste100.com.br'],
    ['4', 'Daniel J S', '(11) 96269-6757', 'danieltisantos@gmail.com'],
]


def valor_ja_existe(workbook, value):
    for row in workbook.active.iter_rows(values_only=True):
        if value in row:
            return True
    return False


workbook = Workbook()
workState = workbook.active

for indexRow, row in enumerate(lista, start=1):
    for indexColum, valueColum in enumerate(row, start=1):
        print(f"Linha {indexRow}, Coluna {indexColum}: {valueColum}")
        if valor_ja_existe(workbook, valueColum):
            print(f"Valor '{valueColum}' já existe na planilha. Pulando gravação.")
        else:
            workState.cell(row=indexRow, column=indexColum, value=valueColum)
workState.title = 'Contatos'
workbook.save('/home/zenxbr/PycharmProjects/PythonProject/treiner/automacao-web/contatos.xlsx')
# Pandas exec

arquivo = '/home/zenxbr/PycharmProjects/PythonProject/treiner/automacao-web/contatos.xlsx'
planilha = pd.read_excel(arquivo)
print(planilha)
