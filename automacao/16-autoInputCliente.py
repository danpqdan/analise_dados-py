import time

import pandas as pd
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

lista = [
    ['nome', 'telefone', 'email', 'estado_civil'],
    ['Wilson', '(11)00000-0000', 'wilson@teste.com', 'Casado'],
    ['Jose da Silva', '(11) 11111-1111', 'jose@teste100.com.br', 'Divorciado'],
    ['Pedro Marcelo', '(11) 22222-2222', 'pedro@teste100.com.br', 'Solteiro'],
    ['Daniel J S', '(11) 96269-6757', 'danieltisantos@gmail.com', 'Solteiro'],
]

workbook = Workbook()
planilha = workbook.active


def valorExistente(workbook, value):
    for row in workbook.active.iter_rows(values_only=True):
        if value in row:
            return True
        return False


for indexRow, row in enumerate(lista, start=1):
    for indexColum, valueColum in enumerate(row, start=1):
        print(f"Linha {indexRow}, Coluna {indexColum}: {valueColum}")
        if valorExistente(workbook, valueColum):
            print(f"Valor '{valueColum}' já existe na planilha. Pulando gravação.")
        else:
            planilha.cell(row=indexRow, column=indexColum, value=valueColum)
planilha.title = 'clientes2'
workbook.save('/home/zenxbr/PycharmProjects/PythonProject/treiner/automacao-web/clientes2.xlsx')

driver = webdriver.Firefox()
arquivo = '/home/zenxbr/PycharmProjects/PythonProject/treiner/automacao-web/clientes2.xlsx'
tabela = pd.read_excel(arquivo)
for index, row in tabela.iterrows():
    print(index)
    print(row["nome"])
    print(row["telefone"])
    print(row["email"])
    print(row["estado_civil"])
    nome = row["nome"]
    telefone = row["telefone"]
    telefoneFormatado = telefone_limpo = telefone.replace("(", "").replace(")", "").replace("-", "").replace(" ", "")
    email = row["email"]
    estado_civil = row["estado_civil"]
    driver.get("https://www.bgmax.com.br/automacao2")
    driver.maximize_window()

    # Preencher Nome
    driver.find_element(By.XPATH, '//*[@id="1459668649"]').send_keys(nome)
    # Preencher Telefone
    driver.find_element(By.XPATH, '//*[@id="1362265994"]').send_keys(telefoneFormatado)
    # Preencher Email
    driver.find_element(By.XPATH, '//*[@id="1804883275"]').send_keys(email)
    # Preencher Estado Civil
    x = driver.find_element(By.XPATH, '//*[@id="1587315858"]')
    sel = Select(x)
    sel.select_by_visible_text(estado_civil)
    time.sleep(3)

    # Clicar Submeter
    # driver.find_element(By.XPATH, '//*[@id="1056525453"]').click()
    time.sleep(15)

time.sleep(10)
driver.close()
