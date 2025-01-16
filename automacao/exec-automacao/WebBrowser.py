from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import pandas as pd
import time
import pyautogui as py
import unicodedata
import os

site = 'https://www.bgmax.com.br/automacao3'
script_path = os.path.abspath(__file__)
diretorioScript = os.path.dirname(script_path)
caminho_index = os.path.join(diretorioScript, 'index.html')
def loadWorkBook():
    if os.path.exists(excel):
        workbook = load_workbook(excel)
        planilha = workbook.active
    else:
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = 'Cadastro-Clientes'
        planilha.append(['Nome', 'Telefone', 'Email', 'Estado_Civil', 'Pratica_Esporte'])
    return workbook, planilha

def salvar_dadosExel(dados):
    workbook, planilha = loadWorkBook()
    for row in dados:
            planilha.append(row)
    workbook.save(f'{diretorioScript}/{nomeDoArquivo}') 

def adicionar_novo_cliente():
    nome = input("Digite o nome: ")
    telefone = input("Digite o telefone (sem caracteres especiais): ")
    email = input("Digite o email: ")
    estado_civil = input("Digite o estado civil (Solteiro/Casado): ")
    pratica_esporte = input("Pratica esporte? (S/N): ")
    cliente = [[nome, telefone, email, estado_civil, pratica_esporte]]
    salvar_dadosExel(dados=cliente)
    return cliente

def iniciar_driver():
    driver = webdriver.Firefox()
    driver.implicitly_wait(36000)
    driver.minimize_window()
    return driver

def formatar_texto(texto):
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    texto = texto.strip().lower()
    if texto in ['s', 'si', 'sm']:
        return 'Sim'
    elif texto in ['n', 'no', 'na']:
        return 'Nao'
    
    return texto

def formatar_estado_civil(estado_civil):
    estado_civil = estado_civil.strip().lower()
    if estado_civil in ['solteiro', 'solteira']:
        return 'Solteiro'
    elif estado_civil in ['casado', 'casada']:
        return 'Casado'
    else:
        print("Estado Civil inválido. Por favor, insira 'Solteiro' ou 'Casado'.")
        return None
    
def formatar_telefone(telefone):
    telefone = telefone.replace("(", "").replace(")", "").replace("-", "").replace(" ", "")
    if telefone.startswith('55'):
        telefone = telefone[2:]
    return telefone

def resgatar_dadosExel(arquivo):
    try:
        planilha = pd.read_excel(arquivo)
        linhas = []
        for index, row in planilha.iterrows():
            linha = {
            "Index": index,
            "Nome": row["Nome"],
            "Telefone": row["Telefone"],
            "Email": row["Email"],
            "Estado_Civil": formatar_estado_civil(row["Estado_Civil"]),
            "Pratica_Esporte": formatar_texto(row["Pratica_Esporte"])
            }
            linhas.append(linha)
        return linhas
    except Exception as e:
        print(f"Erro ao processar o Excel: {e}")

def teste_realizar_cadastro(driver: WebDriver, row: pd.Series):
    actions = ActionChains(driver)
    driver.maximize_window()
    driver.get(f'file://{caminho_index}')
    elemente = driver.find_element(By.XPATH, '//input[@id="nome" and @name="nome"]')
    actions.move_to_element(elemente).click().send_keys(row["Nome"])
    actions.move_to_element(driver.find_element(By.ID, 'telefone')).click().send_keys(row["Telefone"]).perform()
    actions.move_to_element(driver.find_element(By.ID, 'email')).click().send_keys(row["Email"]).perform()
    value = row["Estado_Civil"].lower()
    checkbox_xpath = f"//input[@id='estado_civil' and @value='{value}']"
    driver.find_element(By.XPATH, checkbox_xpath).click()
    value = row["Pratica_Esporte"].lower()
    checkbox_xpath = f"//input[@id='pratica_esporte' and @value='{value}']"
    driver.find_element(By.XPATH, checkbox_xpath).click()
    driver.find_element(By.ID, 'submit').click()
    actions.perform()
    time.sleep(2)

def realizar_cadastro(driver: WebDriver, row: pd.Series):    
    driver.maximize_window()
    driver.find_element(By.LINK_TEXT, 'BGMAX').click()
    driver.switch_to.window(driver.window_handles[1])
    time.sleep(5)
    actions = ActionChains(driver)
    nome = driver.find_element(By.ID, "1459668649")
    actions.move_to_element(nome).click().send_keys(row["Nome"]).perform()
        
    telefone = driver.find_element(By.ID, "1362265994")
    actions.move_to_element(telefone).click().send_keys(row['Telefone']).perform()
    
    email = driver.find_element(By.ID, "1804883275")
    actions.move_to_element(email).click().send_keys(row['Email']).perform()
    
    value = row["Estado_Civil"].capitalize()
    xpath = f'//*[@class="checkable-input" and @name="dmform-3" and normalize-space(@value)="{value}"]'
    element = driver.find_element(By.XPATH, xpath)
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    actions.move_to_element(element).click().perform()
    
    value = row["Pratica_Esporte"].capitalize()
    if value == 'Nao':
        value = 'Não'
    xpath = f'//*[@class="checkable-input" and @name="dmform-62217" and normalize-space(@value)="{value}"]'
    element = driver.find_element(By.XPATH, xpath)
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    actions.move_to_element(element).click().perform()
    time.sleep(2)

    xpath = '//*[@id="1056525453" and @name="submit"]'
    element = driver.find_element(By.XPATH, xpath)
    driver.execute_script("arguments[0].scrollIntoView(true);", element)
    actions.move_to_element(element).click().perform()

    if verificar_recaptcha(driver):
        print("Resolva o reCAPTCHA manualmente.")
        time.sleep(30)
    driver.close()
    driver.switch_to.window(driver.window_handles[0]) 

def verificar_recaptcha(driver: WebDriver):
    try:
        recaptcha_element = driver.find_element(By.CLASS_NAME, 'g-recaptcha')
        print("Página com reCAPTCHA detectada.")
        return True
    except Exception as e:
        print("reCAPTCHA não encontrado.")
        return False
    
def main(arquivo: str):
    try:
        excelConfirm = input('Deseja adicionar um cliente a planilha teste? [S/N]')
        while excelConfirm.lower() not in ['s', 'n']:
            print('Opção invalida...')
            excelConfirm = input('Deseja adicionar um cliente a planilha teste? [S/N]')
        while excelConfirm.lower() == 's':
            adicionado = adicionar_novo_cliente()
            print('Cliente adicionado com sucesso...', adicionado)
            excelConfirm = input('Deseja adicionar mais um cliente? [S/N]')
        print('Todos os clientes foram adicionados...')
        print(80 * '-')
        print('Iniciando o processo de inserção dos dados...')
        print('informe o caminho do arquivo excel com os dados a serem inseridos...')
        print('Por padrão encontra-se na pasta do repositorio junto do script...')
        dados = resgatar_dadosExel(arquivo=arquivo)
        driver = iniciar_driver()
        ultimos_dados = dados[-2:]
        for planilha in ultimos_dados:
            teste_realizar_cadastro(driver, planilha)
            print('Inserido dois ultimos elementos como teste...')
            print(f'Dados inseridos com sucesso: {planilha}')
        driver.minimize_window()
        confirm = input('Deseja concluir a inseção? [S/N]')
        if confirm.lower() =='s':
            for planilha in dados:
                realizar_cadastro(driver, planilha)
                print(f'Dados inseridos com sucesso: {planilha}')
    except Exception as e:
        print(f'Erro: {e}') 
    finally:
        time.sleep(20)
        driver.quit()    

if __name__ == "__main__":
    confirm = input('''Script para inserção de dados em um site pré definido... 
                        pressione qualquer tecla para continuar... Ou digite "Q" para sair...''')
    nomeDoArquivo = input('Digite o nome do arquivo: exemplo.xlsx')
    excel = f'{diretorioScript}/{nomeDoArquivo}'
    if confirm.lower() != 'q':
        loadWorkBook()
        main(arquivo=excel)
    else:
        print('Fim do script...')    
    